from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

import docx2txt
from langchain_community.document_loaders import UnstructuredExcelLoader
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
FIXTURES_DIR = ROOT / "fixtures"
OUTPUTS_DIR = ROOT / "outputs"


def ensure_dirs() -> None:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)


def remove_if_exists(path: Path) -> None:
    if path.exists():
        path.unlink()


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def make_sample_docx(path: Path) -> None:
    paragraphs = [
        "StarRAG Word 样例",
        "",
        "这是一个用于测试 docx2txt 输出的文档。",
        "第二段包含换行符前的内容。",
        "## 这一行本身长得像 Markdown 标题",
        "- 列表项 A",
        "- 列表项 B",
        "键\t值",
        "地区\t华东",
        "备注：本段后面故意保留空行。",
        "",
        "最后一段。",
    ]
    xml_paragraphs = []
    for paragraph in paragraphs:
        if paragraph == "":
            xml_paragraphs.append("<w:p/>")
            continue
        xml_paragraphs.append(
            "<w:p><w:r><w:t xml:space=\"preserve\">"
            f"{escape(paragraph)}"
            "</w:t></w:r></w:p>"
        )

    document_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<w:document "
        "xmlns:wpc=\"http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas\" "
        "xmlns:mc=\"http://schemas.openxmlformats.org/markup-compatibility/2006\" "
        "xmlns:o=\"urn:schemas-microsoft-com:office:office\" "
        "xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\" "
        "xmlns:m=\"http://schemas.openxmlformats.org/officeDocument/2006/math\" "
        "xmlns:v=\"urn:schemas-microsoft-com:vml\" "
        "xmlns:wp14=\"http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing\" "
        "xmlns:wp=\"http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing\" "
        "xmlns:w10=\"urn:schemas-microsoft-com:office:word\" "
        "xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\" "
        "xmlns:w14=\"http://schemas.microsoft.com/office/word/2010/wordml\" "
        "xmlns:wpg=\"http://schemas.microsoft.com/office/word/2010/wordprocessingGroup\" "
        "xmlns:wpi=\"http://schemas.microsoft.com/office/word/2010/wordprocessingInk\" "
        "xmlns:wne=\"http://schemas.microsoft.com/office/word/2006/wordml\" "
        "xmlns:wps=\"http://schemas.microsoft.com/office/word/2010/wordprocessingShape\" "
        "mc:Ignorable=\"w14 wp14\">"
        "<w:body>"
        f"{''.join(xml_paragraphs)}"
        "<w:sectPr><w:pgSz w:w=\"12240\" w:h=\"15840\"/><w:pgMar w:top=\"1440\" "
        "w:right=\"1440\" w:bottom=\"1440\" w:left=\"1440\" w:header=\"720\" "
        "w:footer=\"720\" w:gutter=\"0\"/></w:sectPr>"
        "</w:body></w:document>"
    )

    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>
"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>
"""

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)


def make_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品表"
    ws.append(["产品", "价格", "库存", "备注"])
    ws.append(["苹果", 5, 100, "当季"])
    ws.append(["香蕉", 3, 80, "促销"])
    ws.append(["橙子", 6, 50, "进口"])
    ws.append(["笔记", None, None, "## 这一格像 Markdown 标题"])
    ws["E1"] = "多行字段"
    ws["E2"] = "第一行\n第二行"

    ws2 = wb.create_sheet("季度汇总")
    ws2["A1"] = "区域"
    ws2["B1"] = "Q1"
    ws2["C1"] = "Q2"
    ws2["A2"] = "华东"
    ws2["B2"] = 120
    ws2["C2"] = 150
    ws2["A3"] = "华南"
    ws2["B3"] = 98
    ws2["C3"] = "=B3+12"

    wb.save(path)


def probe_docx(path: Path) -> dict[str, Any]:
    text = docx2txt.process(str(path))
    return {
        "library": "docx2txt",
        "source": str(path),
        "raw_text": text,
        "repr": repr(text),
        "lines": text.splitlines(),
        "line_count": len(text.splitlines()),
    }


def probe_excel_openpyxl(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    sheets = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append(
            {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "rows": rows,
            }
        )
    return {
        "library": "openpyxl",
        "source": str(path),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def probe_excel_unstructured(path: Path) -> dict[str, Any]:
    try:
        from unstructured.partition.xlsx import partition_xlsx
    except Exception as exc:
        return {
            "library": "unstructured.partition.xlsx",
            "source": str(path),
            "status": "import_error",
            "error_type": type(exc).__name__,
            "error": str(exc),
        }

    try:
        elements = partition_xlsx(filename=str(path))
        parsed = []
        for element in elements:
            metadata = getattr(element, "metadata", None)
            parsed.append(
                {
                    "type": element.__class__.__name__,
                    "text": str(element),
                    "repr": repr(str(element)),
                    "metadata": metadata.to_dict() if metadata and hasattr(metadata, "to_dict") else {},
                }
            )
        return {
            "library": "unstructured.partition.xlsx",
            "source": str(path),
            "status": "ok",
            "element_count": len(parsed),
            "elements": parsed,
        }
    except Exception as exc:
        return {
            "library": "unstructured.partition.xlsx",
            "source": str(path),
            "status": "runtime_error",
            "error_type": type(exc).__name__,
            "error": str(exc),
        }


def probe_excel_langchain_mapping(path: Path) -> dict[str, Any]:
    loader = UnstructuredExcelLoader(str(path))
    docs = loader.load()
    unstructured_probe = probe_excel_unstructured(path)
    elements = unstructured_probe.get("elements", []) if isinstance(unstructured_probe, dict) else []

    return {
        "library": "langchain_community.document_loaders.UnstructuredExcelLoader",
        "source": str(path),
        "loader_mode": "single",
        "document_count": len(docs),
        "documents": [
            {
                "page_content": doc.page_content,
                "page_content_repr": repr(doc.page_content),
                "metadata": doc.metadata,
            }
            for doc in docs
        ],
        "element_mapping": [
            {
                "index": idx,
                "element_type": element.get("type"),
                "element_text": element.get("text"),
                "element_text_repr": element.get("repr"),
                "text_as_html": (element.get("metadata") or {}).get("text_as_html"),
                "page_name": (element.get("metadata") or {}).get("page_name"),
                "page_number": (element.get("metadata") or {}).get("page_number"),
            }
            for idx, element in enumerate(elements)
        ],
        "mapping_note": (
            "In single mode, LangChain joins each unstructured element string with "
            "'\\n\\n' and stores the result in Document.page_content."
        ),
    }


def main() -> None:
    ensure_dirs()

    docx_path = FIXTURES_DIR / "sample_word.docx"
    xlsx_path = FIXTURES_DIR / "sample_excel.xlsx"

    remove_if_exists(docx_path)
    remove_if_exists(xlsx_path)
    remove_if_exists(OUTPUTS_DIR / "word_docx2txt.json")
    remove_if_exists(OUTPUTS_DIR / "excel_openpyxl_dump.json")
    remove_if_exists(OUTPUTS_DIR / "excel_unstructured.json")
    remove_if_exists(OUTPUTS_DIR / "excel_langchain_mapping.json")
    remove_if_exists(OUTPUTS_DIR / "summary.json")

    make_sample_docx(docx_path)
    make_sample_xlsx(xlsx_path)

    docx_result = probe_docx(docx_path)
    excel_openpyxl_result = probe_excel_openpyxl(xlsx_path)
    excel_unstructured_result = probe_excel_unstructured(xlsx_path)
    excel_langchain_mapping_result = probe_excel_langchain_mapping(xlsx_path)

    write_json(OUTPUTS_DIR / "word_docx2txt.json", docx_result)
    write_json(OUTPUTS_DIR / "excel_openpyxl_dump.json", excel_openpyxl_result)
    write_json(OUTPUTS_DIR / "excel_unstructured.json", excel_unstructured_result)
    write_json(OUTPUTS_DIR / "excel_langchain_mapping.json", excel_langchain_mapping_result)
    write_json(
        OUTPUTS_DIR / "summary.json",
        {
            "fixtures": {
                "docx": str(docx_path),
                "xlsx": str(xlsx_path),
            },
            "outputs": {
                "word_docx2txt": str(OUTPUTS_DIR / "word_docx2txt.json"),
                "excel_openpyxl_dump": str(OUTPUTS_DIR / "excel_openpyxl_dump.json"),
                "excel_unstructured": str(OUTPUTS_DIR / "excel_unstructured.json"),
                "excel_langchain_mapping": str(OUTPUTS_DIR / "excel_langchain_mapping.json"),
            },
        },
    )


if __name__ == "__main__":
    main()
