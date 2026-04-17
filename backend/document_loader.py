"""文档加载和分片服务"""
import os
from html import escape
from typing import Dict, List, Any

from openpyxl import load_workbook
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader


class DocumentLoader:
    """文档加载和分片服务"""

    def __init__(self, chunk_size: int = 500, chunk_overlap: int = 50):
        # 保留原有参数以兼容外部调用；默认启用三层滑动窗口分块。
        level_1_size = max(1200, chunk_size * 2)
        level_1_overlap = max(240, chunk_overlap * 2)
        level_2_size = max(600, chunk_size)
        level_2_overlap = max(120, chunk_overlap)
        level_3_size = max(300, chunk_size // 2)
        level_3_overlap = max(60, chunk_overlap // 2)

        self._splitter_level_1 = RecursiveCharacterTextSplitter(
            chunk_size=level_1_size,
            chunk_overlap=level_1_overlap,
            add_start_index=True,
            separators=["\n\n", "\n", "。", "！", "？", "，", "、", " ", ""],
        )
        self._splitter_level_2 = RecursiveCharacterTextSplitter(
            chunk_size=level_2_size,
            chunk_overlap=level_2_overlap,
            add_start_index=True,
            separators=["\n\n", "\n", "。", "！", "？", "，", "、", " ", ""],
        )
        self._splitter_level_3 = RecursiveCharacterTextSplitter(
            chunk_size=level_3_size,
            chunk_overlap=level_3_overlap,
            add_start_index=True,
            separators=["\n\n", "\n", "。", "！", "？", "，", "、", " ", ""],
        )

    @staticmethod
    def _build_chunk_id(filename: str, page_number: int, level: int, index: int) -> str:
        return f"{filename}::p{page_number}::l{level}::{index}"

    @staticmethod
    def _normalize_excel_value(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_headers(row_values: List[Any]) -> List[str]:
        headers: List[str] = []
        seen: dict[str, int] = {}
        for idx, value in enumerate(row_values, 1):
            base = DocumentLoader._normalize_excel_value(value) or f"column_{idx}"
            count = seen.get(base, 0)
            seen[base] = count + 1
            headers.append(base if count == 0 else f"{base}_{count + 1}")
        return headers

    @staticmethod
    def _sheet_chunk_id(filename: str, sheet_name: str) -> str:
        return f"{filename}::sheet::{sheet_name}"

    @staticmethod
    def _row_chunk_id(filename: str, sheet_name: str, row_index: int) -> str:
        return f"{filename}::sheet::{sheet_name}::row::{row_index}"

    @staticmethod
    def _build_row_text(filename: str, sheet_name: str, row_index: int, row_obj: dict[str, str]) -> str:
        lines = [f"文件: {filename}", f"Sheet: {sheet_name}", f"行号: {row_index}"]
        for key, value in row_obj.items():
            lines.append(f"{key}={value}")
        return "\n".join(lines)

    @staticmethod
    def _build_sheet_text(filename: str, sheet_name: str, headers: List[str], row_objects: List[dict[str, str]]) -> str:
        header_text = " | ".join(headers)
        preview_rows = row_objects[:5]
        row_lines = []
        for idx, row in enumerate(preview_rows, 1):
            joined = " | ".join([f"{key}={value}" for key, value in row.items() if value != ""])
            row_lines.append(f"示例行{idx}: {joined}")
        if len(row_objects) > len(preview_rows):
            row_lines.append(f"其余 {len(row_objects) - len(preview_rows)} 行已省略")
        return "\n".join([
            f"文件: {filename}",
            f"Sheet: {sheet_name}",
            f"列: {header_text}",
            f"行数: {len(row_objects)}",
            *row_lines,
        ])

    @staticmethod
    def _build_sheet_html(headers: List[str], row_objects: List[dict[str, str]]) -> str:
        header_html = "".join([f"<th>{escape(header)}</th>" for header in headers])
        body_rows = []
        for row in row_objects:
            body_rows.append(
                "<tr>"
                + "".join([f"<td>{escape(row.get(header, ''))}</td>" for header in headers])
                + "</tr>"
            )
        return f"<table><thead><tr>{header_html}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"

    def _split_page_to_three_levels(
        self,
        text: str,
        base_doc: Dict,
        page_global_chunk_idx: int,
    ) -> List[Dict]:
        if not text:
            return []

        root_chunks: List[Dict] = []
        page_number = int(base_doc.get("page_number", 0))
        filename = base_doc["filename"]

        level_1_docs = self._splitter_level_1.create_documents([text], [base_doc])
        level_1_counter = 0
        level_2_counter = 0
        level_3_counter = 0

        for level_1_doc in level_1_docs:
            level_1_text = (level_1_doc.page_content or "").strip()
            if not level_1_text:
                continue
            level_1_id = self._build_chunk_id(filename, page_number, 1, level_1_counter)
            level_1_counter += 1

            level_1_chunk = {
                **base_doc,
                "text": level_1_text,
                "chunk_id": level_1_id,
                "parent_chunk_id": "",
                "root_chunk_id": level_1_id,
                "chunk_level": 1,
                "chunk_idx": page_global_chunk_idx,
            }
            page_global_chunk_idx += 1
            root_chunks.append(level_1_chunk)

            level_2_docs = self._splitter_level_2.create_documents([level_1_text], [base_doc])
            for level_2_doc in level_2_docs:
                level_2_text = (level_2_doc.page_content or "").strip()
                if not level_2_text:
                    continue
                level_2_id = self._build_chunk_id(filename, page_number, 2, level_2_counter)
                level_2_counter += 1

                level_2_chunk = {
                    **base_doc,
                    "text": level_2_text,
                    "chunk_id": level_2_id,
                    "parent_chunk_id": level_1_id,
                    "root_chunk_id": level_1_id,
                    "chunk_level": 2,
                    "chunk_idx": page_global_chunk_idx,
                }
                page_global_chunk_idx += 1
                root_chunks.append(level_2_chunk)

                level_3_docs = self._splitter_level_3.create_documents([level_2_text], [base_doc])
                for level_3_doc in level_3_docs:
                    level_3_text = (level_3_doc.page_content or "").strip()
                    if not level_3_text:
                        continue
                    level_3_id = self._build_chunk_id(filename, page_number, 3, level_3_counter)
                    level_3_counter += 1
                    root_chunks.append({
                        **base_doc,
                        "text": level_3_text,
                        "chunk_id": level_3_id,
                        "parent_chunk_id": level_2_id,
                        "root_chunk_id": level_1_id,
                        "chunk_level": 3,
                        "chunk_idx": page_global_chunk_idx,
                    })
                    page_global_chunk_idx += 1

        return root_chunks

    def load_excel_workbook(self, file_path: str, filename: str) -> dict[str, list[dict]]:
        workbook = load_workbook(file_path, data_only=False)
        sheet_docs: List[dict] = []
        row_docs: List[dict] = []
        sheet_records: List[dict] = []
        row_records: List[dict] = []

        for sheet_index, sheet in enumerate(workbook.worksheets, 1):
            raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header_idx = None
            for idx, row in enumerate(raw_rows):
                if any(self._normalize_excel_value(value) for value in row):
                    header_idx = idx
                    break
            if header_idx is None:
                continue

            headers = self._normalize_headers(raw_rows[header_idx])
            data_rows = raw_rows[header_idx + 1 :]
            row_objects: List[dict[str, str]] = []

            for row_offset, row in enumerate(data_rows, header_idx + 2):
                values = [self._normalize_excel_value(value) for value in row]
                padded = values + [""] * max(0, len(headers) - len(values))
                row_obj = {header: padded[idx] if idx < len(padded) else "" for idx, header in enumerate(headers)}
                if not any(value for value in row_obj.values()):
                    continue
                row_objects.append(row_obj)

                row_chunk_id = self._row_chunk_id(filename, sheet.title, row_offset)
                row_text = self._build_row_text(filename, sheet.title, row_offset, row_obj)
                row_docs.append(
                    {
                        "text": row_text,
                        "filename": filename,
                        "file_type": "Excel",
                        "file_path": file_path,
                        "page_number": sheet_index,
                        "chunk_idx": row_offset,
                        "chunk_id": row_chunk_id,
                        "parent_chunk_id": self._sheet_chunk_id(filename, sheet.title),
                        "root_chunk_id": self._sheet_chunk_id(filename, sheet.title),
                        "chunk_level": 0,
                        "record_type": "excel_row",
                        "sheet_name": sheet.title,
                        "row_index": row_offset,
                    }
                )
                row_records.append(
                    {
                        "chunk_id": row_chunk_id,
                        "filename": filename,
                        "file_path": file_path,
                        "sheet_name": sheet.title,
                        "sheet_index": sheet_index,
                        "row_index": row_offset,
                        "row_text": row_text,
                        "row_obj": row_obj,
                        "headers": headers,
                        "sheet_chunk_id": self._sheet_chunk_id(filename, sheet.title),
                    }
                )

            sheet_chunk_id = self._sheet_chunk_id(filename, sheet.title)
            sheet_text = self._build_sheet_text(filename, sheet.title, headers, row_objects)
            sheet_html = self._build_sheet_html(headers, row_objects)
            sheet_docs.append(
                {
                    "text": sheet_text,
                    "filename": filename,
                    "file_type": "Excel",
                    "file_path": file_path,
                    "page_number": sheet_index,
                    "chunk_idx": 0,
                    "chunk_id": sheet_chunk_id,
                    "parent_chunk_id": "",
                    "root_chunk_id": sheet_chunk_id,
                    "chunk_level": -1,
                    "record_type": "excel_sheet",
                    "sheet_name": sheet.title,
                    "row_index": 0,
                }
            )
            sheet_records.append(
                {
                    "chunk_id": sheet_chunk_id,
                    "filename": filename,
                    "file_path": file_path,
                    "sheet_name": sheet.title,
                    "sheet_index": sheet_index,
                    "headers": headers,
                    "sheet_text": sheet_text,
                    "sheet_html": sheet_html,
                    "row_count": len(row_objects),
                    "column_count": len(headers),
                }
            )

        return {
            "sheets": sheet_records,
            "rows": row_records,
            "sheet_docs": sheet_docs,
            "row_docs": row_docs,
        }

    def load_document(self, file_path: str, filename: str) -> list[dict]:
        """
        加载单个文档并分片
        :param file_path: 文件路径
        :param filename: 文件名
        :return: 分片后的文档列表
        """
        file_lower = filename.lower()

        if file_lower.endswith(".pdf"):
            doc_type = "PDF"
            loader = PyPDFLoader(file_path)
        elif file_lower.endswith((".docx", ".doc")):
            doc_type = "Word"
            loader = Docx2txtLoader(file_path)
        elif file_lower.endswith((".xlsx", ".xls")):
            excel_data = self.load_excel_workbook(file_path, filename)
            return excel_data.get("row_docs", []) + excel_data.get("sheet_docs", [])
        else:
            raise ValueError(f"不支持的文件类型: {filename}")

        try:
            raw_docs = loader.load()
            documents = []
            page_global_chunk_idx = 0
            for doc in raw_docs:
                base_doc = {
                    "filename": filename,
                    "file_path": file_path,
                    "file_type": doc_type,
                    "page_number": doc.metadata.get("page", 0),
                }
                page_chunks = self._split_page_to_three_levels(
                    text=(doc.page_content or "").strip(),
                    base_doc=base_doc,
                    page_global_chunk_idx=page_global_chunk_idx,
                )
                page_global_chunk_idx += len(page_chunks)
                documents.extend(page_chunks)
            return documents
        except Exception as e:
            raise Exception(f"处理文档失败: {str(e)}")

    def load_documents_from_folder(self, folder_path: str) -> list[dict]:
        """
        从文件夹加载所有文档并分片
        :param folder_path: 文件夹路径
        :return: 所有分片后的文档列表
        """
        all_documents = []

        for filename in os.listdir(folder_path):
            file_lower = filename.lower()
            if not (file_lower.endswith(".pdf") or file_lower.endswith((".docx", ".doc")) or file_lower.endswith((".xlsx", ".xls"))):
                continue

            file_path = os.path.join(folder_path, filename)
            try:
                documents = self.load_document(file_path, filename)
                all_documents.extend(documents)
            except Exception:
                continue

        return all_documents
